import os
import re
from statistics import median

import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def normalize_text(value):
    return str(value).strip().upper() if value is not None else ""


def clean_code(code):
    text = normalize_text(code).replace("–", "-")
    text = re.sub(r"\s+", "", text)
    match = re.search(r"(FPA|SF|SA|SH)[A-Z0-9\-_/]*", text)
    return match.group(0) if match else text


def parse_first_number(text):
    if text is None:
        return None
    cleaned = str(text).replace(",", "")
    match = re.search(r"(-?\d+(?:\.\d+)?)", cleaned)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def extract_currency_values(cells):
    values = []
    for cell in cells:
        if cell is None:
            continue
        text = str(cell)
        for number in re.findall(r"€\s*([0-9]+(?:\.[0-9]+)?)", text):
            try:
                values.append(float(number))
            except ValueError:
                continue

    flattened = []
    for cell in cells:
        if cell is None:
            continue
        flattened.extend(str(cell).split())

    for i, token in enumerate(flattened):
        if token in {"€", "EUR"} and i + 1 < len(flattened):
            n = parse_first_number(flattened[i + 1])
            if n is not None:
                values.append(n)

    return values


def split_ams_codes(raw_value):
    parts = re.split(r"[+,/]", str(raw_value))
    return [clean_code(p) for p in parts if normalize_text(p)]


def find_best_sheet(workbook):
    preferred = "PRODUCTION INPUTS LIST"
    exact = None
    fuzzy = None

    for name in workbook.sheetnames:
        normalized = normalize_text(name)
        if normalized == preferred:
            exact = name
            break
        if all(word in normalized for word in ["PRODUCTION", "INPUT", "LIST"]):
            fuzzy = name

    return exact or fuzzy


def process_ams_update(source_path, excel_path, selected_class):
    data_by_code = {}

    def add_row_data(code, length, c1_price, c2_price, mill_price):
        if code not in data_by_code:
            data_by_code[code] = {"length": [], "class1": [], "class2": [], "mill": []}
        if length is not None:
            data_by_code[code]["length"].append(length)
        if c1_price is not None:
            data_by_code[code]["class1"].append(c1_price)
        if c2_price is not None:
            data_by_code[code]["class2"].append(c2_price)
        if mill_price is not None:
            data_by_code[code]["mill"].append(mill_price)

    def parse_table_like_rows(table_rows):
        if not table_rows:
            return

        header_idx = None
        for i, row in enumerate(table_rows[:10]):
            text = " ".join(normalize_text(c) for c in row if c)
            if "AMS" in text and "CLASS" in text:
                header_idx = i
                break

        header_map = {}
        data_rows = table_rows

        if header_idx is not None:
            headers = [normalize_text(c).replace("\n", " ") for c in table_rows[header_idx]]
            data_rows = table_rows[header_idx + 1:]
            for col, h in enumerate(headers):
                if "AMS" in h and "CODE" in h:
                    header_map["ams"] = col
                if "LENGTH" in h:
                    header_map["length"] = col
                if "MILL" in h:
                    header_map.setdefault("mill_cols", []).append(col)
                if "CLASS 1" in h:
                    header_map.setdefault("c1_cols", []).append(col)
                if "CLASS 2" in h:
                    header_map.setdefault("c2_cols", []).append(col)

        for row in data_rows:
            if not row:
                continue
            row = [str(c).strip() if c else "" for c in row]
            joined = " ".join(row).upper()
            if "AMS CODE" in joined:
                continue

            if "ams" in header_map and header_map["ams"] < len(row):
                ams_code = clean_code(row[header_map["ams"]])
            else:
                ams_code = ""
                for cell in row:
                    candidate = clean_code(cell)
                    if candidate.startswith(("FPA", "SF", "SA", "SH")):
                        ams_code = candidate
                        break

            if not ams_code or not ams_code.startswith(("FPA", "SF", "SA", "SH")):
                continue

            # User-validated format: Length (M) is in the 5th column.
            length = parse_first_number(row[4]) if len(row) > 4 else None
            if not length:
                continue

            def read_price(indexes):
                for idx in indexes:
                    if idx >= len(row):
                        continue
                    p = parse_first_number(row[idx])
                    if p is not None and p > 0:
                        return p
                    if row[idx].strip() in {"€", "EUR"} and idx + 1 < len(row):
                        p2 = parse_first_number(row[idx + 1])
                        if p2 is not None and p2 > 0:
                            return p2
                return None

            mill = read_price(header_map.get("mill_cols", []))
            c1 = read_price(header_map.get("c1_cols", []))
            c2 = read_price(header_map.get("c2_cols", []))

            if c1 is None and c2 is None and mill is None:
                money_values = extract_currency_values(row)
                if len(money_values) >= 1:
                    mill = money_values[0]
                if len(money_values) >= 2:
                    c1 = money_values[1]
                if len(money_values) >= 3:
                    c2 = money_values[2]

            add_row_data(ams_code, length, c1, c2, mill)

    source_ext = os.path.splitext(source_path)[1].lower()
    if source_ext == ".pdf":
        with pdfplumber.open(source_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    parse_table_like_rows(table)
    elif source_ext in {".xlsx", ".xlsm"}:
        src_wb = load_workbook(source_path, data_only=True, keep_vba=True)
        for sheet_name in src_wb.sheetnames:
            src_ws = src_wb[sheet_name]
            rows = [
                [src_ws.cell(row=r, column=c).value for c in range(1, src_ws.max_column + 1)]
                for r in range(1, src_ws.max_row + 1)
            ]
            parse_table_like_rows(rows)
    else:
        raise Exception("Unsupported AMS source format. Use PDF or Excel (.xlsx/.xlsm).")

    if not data_by_code:
        raise Exception("No valid data extracted from AMS source")

    ams_lookup = {}
    for code, values in data_by_code.items():
        if selected_class == "1":
            selected_values = values["class1"] or values["mill"] or values["class2"]
        else:
            selected_values = values["class2"] or values["mill"] or values["class1"]

        selected_price = round(median(selected_values), 2) if selected_values else None
        selected_length = round(median(values["length"]), 3) if values["length"] else None

        if selected_price is not None or selected_length is not None:
            ams_lookup[code] = {"length": selected_length, "price": selected_price}

    wb = load_workbook(excel_path, keep_vba=True)
    sheet_name = find_best_sheet(wb)
    if not sheet_name:
        raise Exception("Production Inputs sheet not found")

    ws = wb[sheet_name]
    header_row = None
    for r in range(1, 16):
        values = [
            str(ws.cell(row=r, column=c).value).upper()
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value
        ]
        joined = " ".join(values)
        if "AMS" in joined and "FAMILY" in joined and "LENGTH" in joined and "COST" in joined:
            header_row = r
            break
    if not header_row:
        raise Exception("Header row not found")

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if not val:
            continue
        clean = str(val).strip().upper().replace("\n", " ")
        if "AMS" in clean and "CODE" in clean:
            headers["AMS"] = col
        elif "AMS" in clean and "LENGTH" in clean:
            headers["AMS_LENGTH"] = col
        elif "AMS" in clean and "COST" in clean:
            headers["AMS_COST"] = col
        elif "FAMILY" in clean:
            headers["FAMILY"] = col

    ams_col = headers.get("AMS")
    ams_length_col = headers.get("AMS_LENGTH")
    ams_cost_col = headers.get("AMS_COST")
    family_col = headers.get("FAMILY")
    if not ams_col or not ams_length_col or not ams_cost_col or not family_col:
        raise Exception(f"Columns not found. Detected: {headers}")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    updated_rows = 0
    for row in range(header_row + 1, ws.max_row + 1):
        family = ws.cell(row=row, column=family_col).value
        ams_length_cell = ws.cell(row=row, column=ams_length_col)
        ams_cost_cell = ws.cell(row=row, column=ams_cost_col)
        ams_length_cell.fill = no_fill
        ams_cost_cell.fill = no_fill

        if str(family).strip().upper() != "ALUMINIUM PARTS":
            continue

        ams_value = ws.cell(row=row, column=ams_col).value
        if not ams_value:
            continue

        codes = split_ams_codes(ams_value)
        total_price = 0.0
        matched_lengths = []

        for code in codes:
            info = ams_lookup.get(code)
            if info and info.get("price") is not None:
                total_price += info["price"]
                if info.get("length") is not None:
                    matched_lengths.append(info["length"])

        changed = False
        if matched_lengths:
            ams_length_cell.value = matched_lengths[0]
            ams_length_cell.fill = yellow_fill
            changed = True
        if total_price > 0:
            ams_cost_cell.value = round(total_price, 2)
            ams_cost_cell.fill = yellow_fill
            changed = True
        if changed:
            updated_rows += 1

    base, ext = os.path.splitext(excel_path)
    output_file = base + "_updated" + ext
    wb.save(output_file)
    return output_file, updated_rows
